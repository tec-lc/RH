import csv
from openpyxl import load_workbook

def csv_para_xlsm(arquivo_csv, arquivo_xlsm, sheet_name=None):
    # Carrega a planilha XLSM preservando VBA
    wb = load_workbook(arquivo_xlsm, keep_vba=True)

    # Seleciona a aba
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # -----------------------------
    # 1. Limpa todas as linhas abaixo do cabeçalho (linha 1)
    # -----------------------------
    max_linha = ws.max_row
    if max_linha > 1:
        ws.delete_rows(2, max_linha - 1)

    # -----------------------------
    # 2. Lê o CSV
    # -----------------------------
    with open(arquivo_csv, 'r', encoding='utf-8', newline='') as f:
        reader = csv.reader(f, delimiter=';', quotechar='"')
        linhas = list(reader)

    # -----------------------------
    # 3. Insere os dados a partir da linha 2
    # -----------------------------
    linha_excel = 2
    for row in linhas:
        coluna_excel = 1
        for cell in row:
            ws.cell(row=linha_excel, column=coluna_excel, value=cell)
            coluna_excel += 1
        linha_excel += 1

    # -----------------------------
    # 4. Salva mantendo macros
    # -----------------------------
    wb.save(arquivo_xlsm)
    print("Importação concluída com sucesso!")

# --------------------------
# Exemplo de uso:
# --------------------------

csv_para_xlsm(
    arquivo_csv='downloads/planilha.csv',
    arquivo_xlsm='downloads/modelo.xlsm'
    #sheet_name='template'  # opcional
)
