#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os
from typing import Any


def php_escape(value: Any) -> str:
    """
    Converte um valor Python em uma string PHP segura:
    - None -> null (sem aspas)
    - outros -> string com aspas simples, escapada
    """
    if value is None:
        return "null"

    s = str(value)

    # Escapa barras invertidas e aspas simples
    s = s.replace("\\", "\\\\").replace("'", "\\'")
    # Escapa quebras de linha
    s = s.replace("\r\n", "\\n").replace("\n", "\\n").replace("\r", "\\n")

    return f"'{s}'"


def xlsm_to_php_array(xlsm_path: str) -> str:
    # Carrega o arquivo XLSM
    wb = load_workbook(xlsm_path, data_only=True, keep_vba=True)

    php_lines = []
    php_lines.append("<?php")
    php_lines.append("$val2 = [")

    first_sheet = True

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue  # aba vazia

        header = rows[0]
        data_rows = rows[1:]

        col_data = {}

        for col_idx, col_name in enumerate(header):
            if col_name is None:
                continue  # ignora colunas sem título

            col_values = []
            for row_idx, row in enumerate(data_rows):
                value = row[col_idx] if col_idx < len(row) else None
                col_values.append(value)

            col_data[str(col_name)] = col_values

        if not col_data:
            continue

        if not first_sheet:
            php_lines[-1] += ","
        first_sheet = False

        php_lines.append(f"    '{sheet_name}' => [")

        col_count = len(col_data)
        for i, (col_name, values) in enumerate(col_data.items()):
            php_lines.append(f"        '{col_name}' => [")
            for idx, val in enumerate(values):
                php_val = php_escape(val)
                php_lines.append(f"            {idx} => {php_val},")
            php_lines.append("        ]" + ("," if i < col_count - 1 else ""))

        php_lines.append("    ]")

    php_lines.append("];")
    php_lines.append("")

    return "\n".join(php_lines)


def main():
    # -----------------------------------------------------
    # PARÂMETROS DEFINIDOS INTERNAMENTE AQUI:
    planilha = "/var/www/html/rh/upload/planilha.xlsm"
    script_php = "/var/www/html/rh/planilhas/tempPy.php"
    # -----------------------------------------------------

    # Valida o arquivo XLSM
    if not os.path.isfile(planilha):
        print(f"ERRO: Arquivo XLSM não encontrado: {planilha}")
        return

    # Gera PHP
    try:
        php_code = xlsm_to_php_array(planilha)
    except Exception as e:
        print(f"ERRO ao processar o XLSM: {e}")
        return

    # Garante diretório de saída
    out_dir = os.path.dirname(os.path.abspath(script_php))
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    try:
        with open(script_php, "w", encoding="utf-8") as f:
            f.write(php_code)
    except Exception as e:
        print(f"ERRO ao salvar arquivo PHP: {e}")
        return

    print(f"OK! Arquivo PHP gerado em: {script_php}")


if __name__ == "__main__":
    main()


#adicione a funcionalidade de armazenar os nomes das planilha e colunas importadas em outra variavel $ind que ficara no mesmo arquivo da variavel $val
#modifique o sistema de armazenamento para armazenar os nomes das planilhas e colunas e valores juntos assim

#$ind=[
#    'planilha'=>[//armazena nomes da planilha
#        0=>'nome_planilha'
#    ]
#]

